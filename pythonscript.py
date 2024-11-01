import pandas as pd
import os
from openpyxl.styles import PatternFill
import sys

# Define campaign groups as variables for easy modification
instagram_facebook_campaigns = [
    "INST / Engagement / COLD - 22.05.2024",
    "INST / Engagement Messages / WARM - 20.05.2024",
    "NL / Fb Lead Form / Ugly hair / Lookalike (UA, 1%) - LTV ALL/ 26/08/2024  Campaign",
    "NL / TRAFF / Inst / 24.07.2024",
    "NL / Tailored leads campaign / Lookalike (UA, 1%) - LTV ALL/ 15/07/2024 Campaign",
    "NL / FB Lead Form / Hair transplant /K+O / M22-50 / Broad  18/10/2024   Campaign",
    "NL / FB Lead Form / Regenera 2 /K+O / M22-50 / Broad  24/10/2024   Campaign",
    "NL / Fb Lead Form / Ugly hair / Odesa / Lookalike (UA, 1%)  - LTV ALL/ 21/10/2024"
]

website_campaigns = [
    "NL / Lead  / CBO / Kyiv Odesa / 26.07.2024",
    "SITE / Cold / ABO / Lead / 29.05.2024",
    "SITE / NLAS / Warm / ABO / 12.02.2024"
]

def find_csv_file():
    """Find CSV file in current directory."""
    csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
    if len(csv_files) == 0:
        print("Error: No CSV file found in the current directory")
        sys.exit(1)
    elif len(csv_files) > 1:
        print("Error: Multiple CSV files found in the directory. Please keep only one CSV file.")
        sys.exit(1)
    return csv_files[0]

def check_missing_campaigns(day_data):
    """Check for missing campaigns for a specific day."""
    present_campaigns = day_data['Campaign name'].unique()
    
    missing_ig_fb = [camp for camp in instagram_facebook_campaigns 
                    if camp not in present_campaigns]
    missing_website = [camp for camp in website_campaigns 
                      if camp not in present_campaigns]
    
    comments = []
    if missing_ig_fb:
        comments.append(f"Missing Instagram/Facebook campaigns: {', '.join(missing_ig_fb)}")
    if missing_website:
        comments.append(f"Missing Website campaigns: {', '.join(missing_website)}")
    
    return ' | '.join(comments) if comments else ''

def process_data(csv_file):
    """Process the CSV data and create the required metrics."""
    try:
        # Read CSV file
        print(f"Processing {csv_file}...")
        df = pd.read_csv(csv_file)
        
        # Initialize result dataframes
        results = []
        
        # Process data day by day
        for day in df['Day'].unique():
            day_data = df[df['Day'] == day]
            
            # Calculate metrics for Instagram & Facebook campaigns
            ig_fb_data = day_data[day_data['Campaign name'].isin(instagram_facebook_campaigns)]
            ig_fb_budget = ig_fb_data['Amount spent (USD)'].sum()
            ig_fb_target_leads = (
                ig_fb_data['Leads'].sum() +
                ig_fb_data['Messaging conversations started'].sum() +
                day_data[day_data['Campaign name'].isin(website_campaigns)]['Messaging conversations started'].sum()
            )
            
            # Calculate metrics for Website campaigns
            site_data = day_data[day_data['Campaign name'].isin(website_campaigns)]
            site_budget = site_data['Amount spent (USD)'].sum()
            site_link_clicks = site_data['Link clicks'].sum()
            site_leads = site_data['Leads'].sum()
            
            # Check for missing campaigns
            comments = check_missing_campaigns(day_data)
            
            # Calculate derived metrics
            cpl_ig = ig_fb_budget / ig_fb_target_leads if ig_fb_target_leads != 0 else 0
            cpc_site = site_budget / site_link_clicks if site_link_clicks != 0 else 0
            cpl_site = site_budget / site_leads if site_leads != 0 else 0
            
            results.append({
                'Date': day,
                'Ins&FB_Budget': ig_fb_budget,
                'Ins&FB_Target_leads': ig_fb_target_leads,
                'Site_Budget': site_budget,
                'Site_Link_Clicks': site_link_clicks,
                'Site_Leads': site_leads,
                'Lead Target': site_leads,
                'CPL IG': cpl_ig,
                'CPC Site': cpc_site,
                'CPL Site': cpl_site,
                'Comments': comments
            })
        
        return pd.DataFrame(results)
    
    except Exception as e:
        print(f"Error processing data: {str(e)}")
        sys.exit(1)

def export_simple_xlsx(df, output_file):
    """Export the simple version of the Excel file with colored cells."""
    try:
        # Select and reorder columns for simple version
        simple_columns = ['Date', 'Ins&FB_Budget', 'Ins&FB_Target_leads', 
                         'Site_Budget', 'Site_Link_Clicks', 'Site_Leads', 
                         'Lead Target', 'Comments']
        simple_df = df[simple_columns].copy()
        
        # Export to Excel
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        simple_df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Define fills
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        # Apply colors
        for row in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
            # Yellow for Ins&FB columns
            worksheet.cell(row=row, column=2).fill = yellow_fill  # Ins&FB_Budget
            worksheet.cell(row=row, column=3).fill = yellow_fill  # Ins&FB_Target_leads
            
            # Green for Site columns
            for col in range(4, 8):  # Site_Budget to Lead Target
                worksheet.cell(row=row, column=col).fill = green_fill
            
            # Red for non-empty Comments
            if worksheet.cell(row=row, column=8).value:  # Comments column
                worksheet.cell(row=row, column=8).fill = red_fill
        
        writer.close()
    except Exception as e:
        print(f"Error exporting simple Excel file: {str(e)}")
        sys.exit(1)

def export_detailed_xlsx(df, output_file):
    """Export the detailed version of the Excel file with empty columns."""
    try:
        # Create a new dataframe with empty columns
        detailed_df = pd.DataFrame()
        
        # Add columns in the specified order with empty columns in between
        column_mapping = {
            'A': 'Date',
            'C': 'Ins&FB_Budget',
            'E': 'Ins&FB_Target_leads',
            'F': 'CPL IG',
            'J': 'Site_Budget',
            'K': 'Site_Link_Clicks',
            'L': 'CPC Site',
            'M': 'Site_Leads',
            'N': 'CPL Site',
            'O': 'Lead Target',
            'R': 'Comments'
        }
        
        # Create empty DataFrame with all columns
        all_columns = [''] * 18  # 18 columns (A through R)
        for col_letter, col_name in column_mapping.items():
            col_index = ord(col_letter) - ord('A')
            all_columns[col_index] = col_name
        
        detailed_df = pd.DataFrame(columns=all_columns)
        
        # Fill in the data
        for col_letter, col_name in column_mapping.items():
            col_index = ord(col_letter) - ord('A')
            detailed_df.iloc[:, col_index] = df[col_name]
        
        # Export to Excel
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        detailed_df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get the worksheet
        worksheet = writer.sheets['Sheet1']
        
        # Apply red fill to non-empty Comments cells
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        comments_col = 'R'
        for row in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
            if worksheet.cell(row=row, column=ord(comments_col)-ord('A')+1).value:
                worksheet.cell(row=row, column=ord(comments_col)-ord('A')+1).fill = red_fill
        
        writer.close()
    except Exception as e:
        print(f"Error exporting detailed Excel file: {str(e)}")
        sys.exit(1)

def main():
    try:
        # Find CSV file
        csv_file = find_csv_file()
        
        # Process data
        print("Running the script...")
        results_df = process_data(csv_file)
        
        # Define output filenames
        simple_output = 'simple_report.xlsx'
        detailed_output = 'detailed_report.xlsx'
        
        # Export Excel files
        export_simple_xlsx(results_df, simple_output)
        export_detailed_xlsx(results_df, detailed_output)
        
        # Print success message with file paths
        print(f"\nExported files successfully:")
        print(f"Simple report: {os.path.abspath(simple_output)}")
        print(f"Detailed report: {os.path.abspath(detailed_output)}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()